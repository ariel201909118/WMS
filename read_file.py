import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Side, Border, alignment, PatternFill
from duplicate import highlight_duplicate

def read(filename, container_no, reference_no):
    df = pd.read_excel(filename, sheet_name='PO&Item Details')
    data = df[df['Trailer'] == container_no]
    
    config_ = [
        [3 ,'Min', ''],[4, 'Max', ''], [5, 'Cycle Count', ''], [6, 'ReorderQty', ''], [7, 'Inventory Method', ''],
        [8, 'Temperature', ''], [9, 'Cost', ''], [10, 'UPC', ''], [11, 'Track Lot','FALSE'], [12, 'Track Serial', 'FALSE'],
        [13, 'Track ExpDate', 'FALSE'], [14, 'Primary Unit of Measure', 'Carton'], [15, 'Packaging Unit', 'Pallet'],
        [16, 'Packing UOM Qty', 1], [17, 'Length J', 1], [18, 'Width', 1], [19, 'Height', 1], [20, 'Weight', ''],
        [21, 'Qualifiers', ''], [22, 'Storage Setup', ''], [23, 'Variable Setup', ''], [24, 'NMFC#', ''],
        [25, 'Lot Number Required',  'FALSE'], [26, 'Serial Number Required', 'FALSE'], 
        [27, 'Serial Number (must be unique)', 'FALSE'], [28, 'Exp Date Req', 'FALSE'], [29, 'E1ble Cost', 'FALSE'],
        [30, 'Cost Required', 'FALSE'], [31, 'IsHazMat', ''], [32, 'HazMatID', ''], [33, 'HazMatShipping1me',''],
        [34, 'HazMatHazardClass', ''], [35, 'HazMatPackingGroup', ''], [36, 'HazMatFlashPoint', ''], 
        [37, 'HazMatLabelCode', ''], [38, 'HazMatFlag', ''], [39, 'Image URL', ''],[40, 'StorageCountScriptTemplateID', ''],
        [41, 'StorageRates', ''], [42, 'OutboundMobileSerializationBehavior', ''], [43, 'Price', ''], [44, 'TotalQty', ''],
        [45, 'UnitType', '']
    ]

    if data['Trailer'].shape[0] != 0:
        item = data[['WMT ItemNumber', 'DESCRIPTION', 'DCPO']].drop_duplicates(subset=['WMT ItemNumber'])
        for x in config_:
            item.insert(x[0], x[1], x[2])
        item.to_csv(f'Items {container_no}.txt', index=None, header=None, sep='\t', float_format='%.0f')#creates the items file to import to WMS
        #Create the file to print
        
        paper = data[['Trailer', 'DCPO', 'DESCRIPTION', 'WMT ItemNumber', 'Ctns']]
        paper = paper.rename(columns={'WMT ItemNumber': 'Sku'})
        paper_ = ((5,'T',''), (6, 'H', ''), (7, '', ''), (8, 'TOTAL',''))
        for x in paper_:
            paper.insert(x[0], x[1], x[2])
        
        header = paper.columns.values
        
        
        wb = Workbook()
        ws1 = wb.create_sheet("Reference#")
        sheet = wb.active
        sheet.title = 'Items'
        
        def style(file, col, _width, col_pos, cell_width, col_height=35, sheet=wb.active):
            x = len(col)
                
            thin_border = Border(
                left=Side(style='thin'),right=Side(style='thin'),
                top=Side(style='thin'),bottom=Side(style='thin')
                )
            for cont in range(2, x + 2):
                sheet.cell(row=cont, column=col_pos).value = col[cont - 2]
                sheet.cell(row=cont, column=col_pos).border = thin_border
                sheet.cell(row=cont, column=col_pos).alignment = alignment.Alignment(horizontal='left')
                sheet.row_dimensions[cont].height = col_height
                sheet.column_dimensions[cell_width].width = _width
            wb.save(f'{file}.xlsx')
        
        
        for c in range(1,len(header)+1):    
            sheet.cell(row=1, column=c).value = header[c-1]
        wb.save(f'{container_no}.xlsx')
        
        
        style(container_no, paper['Trailer'].values, 12, 1, 'A', sheet=wb.active)
        style(container_no, paper['DCPO'].values, 11.5, 2, 'B', sheet=wb.active) 
        style(container_no, paper['DESCRIPTION'].values, 31, 3,'C', sheet=wb.active)
        style(container_no, paper['Sku'].values, 9, 4, 'D', sheet=wb.active)
        style(container_no, paper['Ctns'].values, 6, 5, 'E', sheet=wb.active) 
        style(container_no, paper['T'].values, 6, 6, 'F', sheet=wb.active) 
        style(container_no, paper['H'].values, 6, 7, 'G', sheet=wb.active) 
        style(container_no, paper[''].values, 44, 8, 'H', sheet=wb.active) 
        style(container_no, paper['TOTAL'].values, 8, 9, 'I', sheet=wb.active)
        
        compare = highlight_duplicate(paper['Sku'].values.tolist())
        
        if compare != []:
            for x in compare:
                sheet.cell(row=x+2, column=4).fill = PatternFill(start_color='FFEE08', end_color='00FF9900', fill_type='solid')
            wb.save(f'{container_no}.xlsx')
        
        receipt_len = data.shape[0]
        ref = []
        for x in range(receipt_len):
            ref.append('XD00000'+ str(int(reference_no) + x))    

        nw_df = data[['DCPO']]
        receipt_config = [
            [0, 'Ref #', ref], [2, 'ShipCarrier', ''], [3, 'Description', paper['DESCRIPTION']],
            [4, 'Sku', paper['Sku']], [5, 'SKU Quantity', 1], [6, 'lot #', ''], [7, 'Serial#', ''], [8, 'Expiration Date', ''],
            [9, 'LocationField1', 'STAGE 1'], [10, 'LocationField2', ''], [11, 'LocationField3', ''], [12, 'LocationField4',''],
            [13, 'Cost', ''], [14, 'Var UOM Avg', ''], [15, 'Receipt Advice Number', paper['Trailer']], 
            [16, 'TrackingNumber', paper['Trailer']], [17, 'SupplierCompanyName', ''], [18, 'MULabel', 'SYSTEMSET'],
            [19, 'MUType', 'Pallet'], [20, 'MUTypeMULength', 1], [21, 'MUWidth', 1], [22, 'MUHeight', 1],
            [23, 'MUWeight', 1], [24, 'Receiver Transaction Saved Element', ''], [25, 'CreateMulipleMUs', 'CreateMulipleMUs:TRUE'],
            [26, 'Catch Weight', '']                                                                                    
        ]
        for x in receipt_config:
            nw_df.insert(x[0], x[1], x[2])
        nw_df.to_csv(f'Receipt {container_no}.txt', index=None, header=None, sep='\t', float_format='%.0f')#creates the receipt file to import to WMS
        
        paper_nw = nw_df[['Ref #', 'DCPO', 'Sku','Catch Weight']]
        print(wb.sheetnames)
        wb.active = 1
        sheet = wb.active
        print(sheet)
        style((container_no),paper['Trailer'].values, 12, 1, 'A', 20, sheet=wb.active)
        style((container_no),paper_nw['Ref #'].values, 13, 2, 'B', 20, sheet=wb.active)
        style((container_no),paper_nw['DCPO'].values, 11.5, 3, 'C', 20, sheet=wb.active) 
        style((container_no),paper_nw['Sku'].values, 9, 4, 'D', 20, sheet=wb.active)
        style((container_no),paper_nw['Catch Weight'].values, 44, 5, 'E',20, sheet=wb.active)
        
        header1 = ['Trailer', 'Ref#', 'DCPO', 'SKU', '']
        for c in range(1,len(header1)+1):    
            sheet.cell(row=1, column=c).value = header1[c-1]
        wb.save(f'{container_no}.xlsx')
        
        compare = highlight_duplicate(paper_nw['Sku'].values.tolist())
        if compare != []:
            for x in compare:
                sheet.cell(row=x+2, column=4).fill = PatternFill(start_color='FFEE08', end_color='00FF9900', fill_type='solid')
            wb.save(f'{container_no}.xlsx')
            
        return(nw_df['Ref #'].iat[-1])
        
    else:
        return('Container Not found')   